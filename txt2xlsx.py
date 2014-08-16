#!/usr/bin/env python3

"""
Author              : Luke Snow
Date Created        : 8/16/14
Date Last Modified  : 8/16/14

Description:
  This is a simple module that will convert a text file into an xlsx file.

Version:
  1.0.0
"""

# Taken from stackoverflow:
def is_number(s):
  """This function checks if a string can be converted to a number"""
  try:
    float(s)
    return True
  except ValueError:
    return False

def main(input_file, output_file):
  """This is a simple function that will convert a text file into an xlsx file."""
  # Import modules
  import csv, os
  from openpyxl import Workbook

  # Open text file with csv reader.
  in_fl_obj = open(input_file)
  csv_obj=csv.reader(in_fl_obj)

  # Create new excel workbook object.
  wb=Workbook()
  ws=wb.active
  ws.title = os.path.basename(input_file)

  row_idx=1

  # Each row in the text file, write all values on the row to the XLSX file.
  for rows in csv_obj:
    # First, check if row is blank.  If so, write an empty.
    if not rows:
      ws.cell(row=row_idx, column=1).value = ""
    else:
      for col_idx in range(1, len(rows[0].split())+1):
        input_str = rows[0].split()[col_idx-1]
        if is_number(input_str[col_idx-1]):
          ws.cell(row=row_idx, column=col_idx).value = float(input_str)
        else: 
          ws.cell(row=row_idx, column=col_idx).value = input_str
    row_idx += 1

  wb.save(filename = output_file)

if __name__ == "__main__":
  import sys
  if len(sys.argv) != 3:
    print("Error, wrong number of arguments.")
    print("Correct syntax is: txt2xlsx <input file> <output file>")
    sys.exit()
  else:
    main(sys.argv[1], sys.argv[2])

